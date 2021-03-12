import csv
from uuid import UUID

from openpyxl import load_workbook, Workbook

from backend.src.apps.xlsx.serializer import prepare_current_data_to_mongo, prepare_new_data_to_mongo


def parse_xlsx_to_convert_data_to_dict(filepath: str) -> tuple:
    """
    Функция для парсинга данных из xlsx файла.

    Возвращает ключи с названием столбцом таблицы и их значениями в виде строк.
    Данная функция должна быть в репозитории где скачивается файл со статистикой.
    """
    rows = []  # Данные книг
    keys = []  # Шапка документа с названием столбцом.

    file = load_workbook(filename=filepath)  # Открытие документа
    sheet = file.active  # Выбор активного листа

    for index, row in enumerate(sheet.rows):
        if index == 0:
            keys = [x.value for x in row]
            continue

        # Преобразуем данные внутри строки в список
        rows.append([x.value for x in row])

    return keys, rows


def convert_xlsx_rows_to_dict(keys: list, rows: list) -> list:
    """
    Преобразование XLSX данных в словари.

    На выходе получаем список словарей в формате [{isbn: ..., hours: ..., author: ...}, {}, {}]
    Данная функция должна быть в репозитории где скачивается файл со статистикой.
    """
    converted_data = []  # Готовая дата

    for row in rows:
        row_to_dict = {}  # Строка XLSX преобразованная в словарь

        # Перебор значений в строке XLSX и присваивание ключа к значению по номеру индекса из списка ключей.
        for index, item in enumerate(row):
            row_to_dict[keys[index]] = item

        converted_data.append(row_to_dict)

    return converted_data


def merge_lines(current_dashboard: list, new_dashboard: list) -> list:
    """
    Функция для слияния двух строк из разных таблицы XLSX необходимо для подготовки данных в MongoDB.

    Данная функция должна быть в репозитории где скачивается файл со статистикой.
    """

    merged_items = []

    for dict_current_data in current_dashboard:
        for key, value in dict_current_data.items():  # Перебор ключей в строке основонго дашборда
            if key == "Title":  # Если ключ Title начинаем поиск по названию внутри нового дашборда

                row_is_find = False  # Начальное состояние поиска

                for dict_new_data in new_dashboard:  # Перебор ключей в строке основонго дашборда

                    if value == dict_new_data['Title']:  # Если значения совпадают собираем все в одну строку.
                        data = {
                            **prepare_current_data_to_mongo(dict_current_data),
                            **prepare_new_data_to_mongo(dict_new_data),
                        }

                        merged_items.append(data)

                        new_dashboard.remove(dict_new_data)  # Удаляем из нового дашборда найденный результат.

                        row_is_find = True  # Переключаем состояние что строка найдена и выходим из дальнейшего поиска.
                        break

                if row_is_find is False:  # Иначе заполняем дефолтными значениями из нового дашборда
                    data = {
                        **prepare_current_data_to_mongo(dict_current_data),
                        **prepare_new_data_to_mongo(),
                    }
                    merged_items.append(data)

    return merged_items


def write_data_to_xlsx(data, path_to_file):
    """Запись подготовленных данных статистики в xlsx"""

    # Создаем файл для записи данных пользователя.
    new_xlsx = Workbook()
    sheet_for_writing = new_xlsx.active

    # index - нужен для нумерации строк. row - список данных для записи.
    for index, row in enumerate(data):

        for item in range(len(row)):
            # Нумерация строк и колон начинается с 1, добавляем к индуксу 1.
            new_row, column = index + 1, item + 1

            value = row[item]
            # openpyxl не может записать UUID, проверяем и преобразуем UUID в str.
            if isinstance(value, UUID):
                value = str(value)

            sheet_for_writing.cell(row=new_row, column=column, value=value)

    new_xlsx.save(path_to_file)


def prepare_data_from_csv_for_xlsx(file_path_to_csv: str):
    """
    Получаем подготовленный к записи в xlsx список с данными из файла CSV.

    Каждый элемента списка - список с данными колонок.
    :param file_path_to_csv: Путь к CSV файлу с данными для записи в xlsx.
    """
    prepared_data_for_xlsx = list()

    with open(file_path_to_csv, 'r', encoding='utf-8') as file:
        csv_file = csv.reader(file)
        for row in csv_file:
            prepared_data = list()  # Список для данных из строки
            for value in row:
                prepared_data.append(value)
            prepared_data_for_xlsx.append(prepared_data)

    return prepared_data_for_xlsx
